"""
更新数据清单: D:\2026-SP\Data_checklist.xlsx
同步当前真实数据状态
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import os

CHECKLIST = r"D:\2026-SP\Data_checklist.xlsx"
CSV = r"D:\2026-SP\Outputs\pinggu_environmental_data.csv"

# 读取已有workbook
wb = openpyxl.load_workbook(CHECKLIST)
ws = wb.active

# 定义颜色
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_font = Font(color="9C6500", bold=True)
red_font = Font(color="9C0006", bold=True)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# 当前状态更新规则
updates = {
    # 维度编号 → (原状态, 新状态, 备注)
    # 1-19: bioclim → 保持"已获取"
    # 20-33: monthly → 保持"已获取"
    # 34-39: GDD → 已在CSV中计算，更新为"已获取(已计算)"
    34: ("可计算", "已获取(CSV计算)", "从tavg月度数据计算，已存入pinggu_environmental_data.csv"),
    35: ("可计算", "已获取(CSV计算)", "从tavg月度数据计算，已存入pinggu_environmental_data.csv"),
    36: ("可计算", "已获取(CSV计算)", "从tavg月度数据计算，已存入pinggu_environmental_data.csv"),
    37: ("可计算", "已获取(CSV计算)", "从tavg月度数据计算，已存入pinggu_environmental_data.csv"),
    38: ("可计算", "已获取(CSV计算)", "从tavg月度数据计算，已存入pinggu_environmental_data.csv"),
    39: ("可计算", "已获取(CSV计算)", "从tavg月度数据计算，已存入pinggu_environmental_data.csv"),
    # 40-42: 极端指数 → CHELSA失败，改用bio5/bio6/bio13代理
    40: ("需手动", "已获取(代理)", "CHELSA逐日下载不可行(2.5TB)，改用WorldClim bio5代理Tmax_95p"),
    41: ("需手动", "已获取(代理)", "CHELSA逐日下载不可行(2.5TB)，改用WorldClim bio6代理Tmin_5p"),
    42: ("需手动", "已获取(代理)", "CHELSA逐日下载不可行(2.5TB)，改用WorldClim bio13代理Prec_95p"),
    # 43: CO2 → 保持"已获取"
    # 44-51: SoilGrids → 原Homolosine文件不可用，已下载WGS84版并提取
    44: ("已获取", "已获取(WGS84)", "5km WGS84版下载+重投影，已提取到CSV"),
    45: ("已获取", "已获取(WGS84)", "5km WGS84版下载+重投影，已提取到CSV"),
    46: ("已获取", "已获取(WGS84)", "5km WGS84版下载+重投影，已提取到CSV"),
    47: ("已获取", "已获取(WGS84)", "5km WGS84版下载+重投影，已提取到CSV"),
    48: ("已获取", "已获取(估算)", "nitrogen下载失败，按SOC/10估算，已提取到CSV"),
    49: ("已获取", "已获取(WGS84)", "5km WGS84版下载+重投影，已提取到CSV"),
    50: ("已获取", "已获取(WGS84)", "5km WGS84版下载+重投影，已提取到CSV"),
    51: ("已获取", "已获取(WGS84)", "5km WGS84版下载+重投影，已提取到CSV"),
    # 52-54: SAT/DUL/LL15 → 中国土壤数据仍缺
    # 55: elevation → 更新为正确文件
    55: ("已获取", "已获取(SRTM90m)", "SRTM tile srtm_60_04.tif，重采样到4.6km网格，已提取到CSV"),
    # 56-57: slope/aspect → 已在CSV中计算
    56: ("可计算", "已获取(CSV计算)", "从SRTM elevation计算，已存入pinggu_environmental_data.csv"),
    57: ("可计算", "已获取(CSV计算)", "从SRTM elevation计算，已存入pinggu_environmental_data.csv"),
}

# 状态对应的颜色和字体
def get_style(status_text):
    if "已获取" in status_text:
        return green_fill, green_font
    elif "可计算" in status_text or "需手动" in status_text:
        return yellow_fill, yellow_font
    elif "模板" in status_text:
        return red_fill, red_font
    else:
        return None, None

# 更新行
for row in range(2, ws.max_row + 1):
    dim_no = ws.cell(row=row, column=1).value
    if dim_no is None:
        continue
    
    try:
        dim_no = int(dim_no)
    except:
        continue
    
    if dim_no in updates:
        old_status, new_status, note = updates[dim_no]
        current_status = ws.cell(row=row, column=6).value
        
        # 更新状态列
        ws.cell(row=row, column=6).value = new_status
        
        # 更新源文件列(如果有变化)
        if dim_no == 55:
            ws.cell(row=row, column=5).value = "Data/SRTM/srtm_60_04.tif"
        if 44 <= dim_no <= 51:
            ws.cell(row=row, column=4).value = "SoilGrids 2.0 (5km WGS84)"
            # 更新文件路径
            prop_map = {44:'sand',45:'silt',46:'clay',47:'soc',48:'nitrogen',49:'phh2o',50:'bdod',51:'cec'}
            if dim_no in prop_map:
                ws.cell(row=row, column=5).value = f"Data/SoilGrids_wgs84/{prop_map[dim_no]}_0-5cm_mean_5000.tif"
        
        # 应用样式
        fill, font = get_style(new_status)
        if fill:
            ws.cell(row=row, column=6).fill = fill
            ws.cell(row=row, column=6).font = font

# 数据来源说明
source_updates = {
    40: "WorldClim 2.1 bio5 (代理)",
    41: "WorldClim 2.1 bio6 (代理)",
    42: "WorldClim 2.1 bio13 (代理)",
}
for dim_no, new_source in source_updates.items():
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == dim_no:
            ws.cell(row=row, column=4).value = new_source
            ws.cell(row=row, column=5).value = "Data/WorldClim (已含在bio变量中)"

# 添加注释sheet
if '变更记录' in wb.sheetnames:
    ws_notes = wb['变更记录']
else:
    ws_notes = wb.create_sheet('变更记录')

ws_notes['A1'] = '日期'
ws_notes['B1'] = '变更内容'
ws_notes['A1'].font = Font(bold=True)
ws_notes['B1'].font = Font(bold=True)

# 找到最后一行
last_row = ws_notes.max_row + 1
notes = [
    ('2026-07-21', '极端气候指数: CHELSA逐日下载不可行→改用WorldClim bio5/bio6/bio13代理'),
    ('2026-07-21', 'SoilGrids: 原Homolosine投影文件不可用→下载5km WGS84版(clay/sand/silt/soc/bdod/cec/phh2o/nitrogen)'),
    ('2026-07-21', 'SRTM: 替换正确tile (srtm_60_04, 115-120E,40-45N覆盖平谷)'),
    ('2026-07-21', 'GDD: 从月度气温计算(1-12月+生长季+全年)，已存入主CSV'),
    ('2026-07-21', 'Slope/Aspect: 从SRTM 90m计算并重采样到4.6km网格'),
    ('2026-07-21', 'Nitrogen: ISRIC下载失败，按SOC/10估算'),
    ('2026-07-21', '主数据文件: D:/2026-SP/Outputs/pinggu_environmental_data.csv (234 rows × 73 cols)'),
]
for i, (date, note) in enumerate(notes):
    ws_notes.cell(row=last_row+i, column=1).value = date
    ws_notes.cell(row=last_row+i, column=2).value = note

# 保存
wb.save(CHECKLIST)
print(f'✅ 清单已更新: {CHECKLIST}')
print(f'   更新了 {len(updates)} 行状态')

# 统计
status_counts = {}
for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=6).value
    if status:
        status_counts[status] = status_counts.get(status, 0) + 1

print(f'\n状态统计:')
for s, c in sorted(status_counts.items()):
    print(f'  {s}: {c}')
print(f'  总计: {sum(status_counts.values())}')

# 验证CSV存在
csv_path = r"D:\2026-SP\Outputs\pinggu_environmental_data.csv"
if os.path.exists(csv_path):
    import pandas as pd
    df = pd.read_csv(csv_path)
    print(f'\n主CSV验证: {csv_path}')
    print(f'  形状: {df.shape}')
    print(f'  大小: {os.path.getsize(csv_path)/1024:.0f} KB')
